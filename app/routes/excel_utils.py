import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def crear_excel_estilizado(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = [
        "Fecha",
        "Usuario",
        "Tipo",
        "Descripción",
        "Cantidad",
        "Precio Unitario",
        "Precio Total",
    ]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="232b3b")
    cell_fill_ingreso = PatternFill("solid", fgColor="e6f7e6")
    cell_fill_egreso = PatternFill("solid", fgColor="fff0e6")
    border = Border(
        left=Side(style="thin", color="7e6520"),
        right=Side(style="thin", color="7e6520"),
        top=Side(style="thin", color="7e6520"),
        bottom=Side(style="thin", color="7e6520"),
    )

    # Encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_ingreso = 0
    total_egreso = 0

    # Datos
    for row in resultados:
        ws.append(row)
        tipo = str(row[2]).lower()
        precio_total = float(row[6]) if row[6] else 0
        if tipo == "ingreso":
            total_ingreso += precio_total
        elif tipo == "egreso":
            total_egreso += precio_total

    # Estilos a filas de datos
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7), start=2
    ):
        tipo = ws.cell(row=idx, column=3).value
        fill = cell_fill_ingreso if tipo == "Ingreso" else cell_fill_egreso
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Ajusta ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Fila de totales
    ws.append([])
    ws.append(["", "", "", "", "", "Total Ingreso:", total_ingreso])
    ws.append(["", "", "", "", "", "Total Egreso:", total_egreso])
    ws.append(["", "", "", "", "", "Total del Día:", total_ingreso - total_egreso])

    for i in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=i, column=6).font = Font(bold=True, color="232b3b")
        ws.cell(row=i, column=7).font = Font(bold=True, color="232b3b")
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="F0B000")
        ws.cell(row=i, column=7).fill = PatternFill("solid", fgColor="F0B000")
        ws.cell(row=i, column=6).border = border
        ws.cell(row=i, column=7).border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
